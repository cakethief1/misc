''' Requirements: openpyxl, Pillow '''

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from datetime import date, timedelta


today = date.today()
tomorrow = today + timedelta(days = 1)


chilled_manifests = {


					'CJ LANG': 			'CJ LANG CHILLED ESTIMATES MANIFEST',
					'AF BLAKEMORE': 	'AF BLAKEMORE CHILLED ESTIMATES MANIFEST',
					'TALBOT GREEN': 	'TALBOT GREEN CHILLED ESTIMATES MANIFEST',
					'JAMES HALL': 		'JAMES HALL CHILLED ESTIMATES MANIFEST',
					'APPLEBY WESTWARD': 'APPLEBY WESTWARD CHILLED ESTIMATES MANIFEST',
					'HENDERSONS': 		'HENDERSONS CHILLED ESTIMATES MANIFEST'


					}


output_directory = 'C:\\Users\\bens\\Desktop\\test\\'

template_file =  'C:\\Users\\bens\\Desktop\\test\\' + 'template' + '.xlsx'

img = Image('C:\\Users\\bens\\Desktop\\test\\spar_logo.jpg')


for manifest_name in chilled_manifests.values():
	output_file = '{}'.format(output_directory) + '{}'.format(manifest_name) + ' ' + '{}'.format(tomorrow.strftime("%d.%m.%y")) + '.xlsx'
	wb = load_workbook(template_file)
	ws = wb.active
	ws['F13'] = tomorrow.strftime("%d.%m.%y")
	ws.add_image(img, 'A1')
	wb.save(output_file)
	print('File: "{}'.format(manifest_name) + ' {}'.format(tomorrow.strftime("%d.%m.%y") + '.xlsx"' + ' ' + ' created in' + ' ' + '"{}"'.format(output_directory)))